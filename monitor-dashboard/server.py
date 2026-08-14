"""CPS达人全链路操盘2.0 看板后端服务"""

import asyncio
import csv
import json
import math
import os
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any, Optional
from urllib.parse import unquote

import httpx
from fastapi import FastAPI, Query, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware

import config
from openpyxl import load_workbook

app = FastAPI(title="CPS达人运营看板", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _load_json(path: str) -> dict:
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: load {path}: {e}")
    return {}


def _num(v: Any) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


class DataLoader:
    def __init__(self):
        self.report_path = config.REPORT_DATA_PATH
        self._report_cache: dict = {}
        self._report_mtime: float = -1
        self._content_cache: list[dict] = []
        self._content_cache_mtimes: tuple[float, float, float] | None = None
        self._traffic_cache: list[dict] = []
        self._traffic_mtime: float = -1

    @staticmethod
    def _investment_date(value: Any) -> Optional[date]:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value or "").strip().replace("/", ".").replace("-", ".")
        for fmt in ("%Y.%m.%d", "%Y.%m", "%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
        return None

    def get_roi_analysis(self, tracking_days: int = 30) -> dict:
        investments = _load_json(config.ROI_DATA_PATH).get("items", [])
        facts = self._content_catalog()
        sales_by_name = {}
        for row in self.get_report().get("creators", []) or []:
            for name in str(row.get("bilibili_accounts") or row.get("creator_display") or "").split("/"):
                if name.strip(): sales_by_name[name.strip()] = row
        result = []
        for index, item in enumerate(investments, 1):
            name = str(item.get("creator_name") or "").strip()
            start = self._investment_date(item.get("investment_date"))
            end = start + timedelta(days=tracking_days - 1) if start else None
            matched = [row for row in facts if row.get("creator_name") == name and start
                and (row_date := self._date(row.get("date"))) and start <= row_date <= end]
            sales = sales_by_name.get(name, {})
            cost = item.get("investment_amount")
            commission = _num(sales.get("commission_amount"))
            result.append({**item, "index": index, "tracking_days": tracking_days,
                "period_start": start.isoformat() if start else "", "period_end": end.isoformat() if end else "",
                "content_count": len(matched),
                "play_count": sum(int(_num(row.get("play_count"))) for row in matched),
                "interaction_count": sum(int(_num(row.get("interaction_count"))) for row in matched),
                "blue_link_count": sum(int(_num(row.get("blue_link_count"))) for row in matched),
                "conversion_count": int(_num(sales.get("sales_quantity"))) if sales else None,
                "attributed_commission": commission if sales else None,
                "roi": round(commission / float(cost), 2) if sales and cost and float(cost) > 0 else None,
                "match_status": "matched" if matched or sales else "unmatched"})
        return {"items": result, "count": len(result), "tracking_days": tracking_days}

    def get_report(self) -> dict:
        try:
            mtime = os.path.getmtime(self.report_path)
        except OSError:
            return {}
        if mtime != self._report_mtime:
            self._report_cache = _load_json(self.report_path)
            self._report_mtime = mtime
        return self._report_cache

    @staticmethod
    def _date(value: Any) -> Optional[date]:
        text = str(value or "")[:10]
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            return None

    def _resolve_period(self, start_date: Optional[str] = None, end_date: Optional[str] = None,
                        days: int = 7) -> tuple[date, date]:
        daily_dates = [self._date(d.get("date")) for d in self.get_report().get("daily", [])]
        available = [d for d in daily_dates if d]
        default_end = max(available) if available else datetime.now().date()
        end = self._date(end_date) or default_end
        start = self._date(start_date) or (end - timedelta(days=max(1, days) - 1))
        if start > end:
            raise ValueError("start_date 不能晚于 end_date")
        return start, end

    @staticmethod
    def _period_label(start: date, end: date) -> tuple[str, str]:
        return start.isoformat(), end.isoformat()

    @staticmethod
    def _previous_period(start: date, end: date) -> tuple[date, date]:
        span = (end - start).days + 1
        previous_end = start - timedelta(days=1)
        return previous_end - timedelta(days=span - 1), previous_end

    def _period_daily(self, start: date, end: date) -> list[dict]:
        rows_by_date: dict[str, dict] = {}
        for row in self.get_report().get("daily", []) or []:
            row_date = self._date(row.get("date"))
            if row_date and start <= row_date <= end:
                rows_by_date[row_date.isoformat()] = dict(row)
        # Traffic-source data is delivered independently from the report JSON.
        # Merge it by date so every dashboard filter uses one consistent period.
        for row in self._traffic_daily():
            row_date = self._date(row.get("traffic_date"))
            if row_date and start <= row_date <= end:
                key = row_date.isoformat()
                rows_by_date.setdefault(key, {"date": key}).update(row)
        return [rows_by_date[key] for key in sorted(rows_by_date)]

    def _traffic_daily(self) -> list[dict]:
        path = config.TRAFFIC_FACTS_PATH
        try:
            mtime = os.path.getmtime(path)
        except OSError:
            return []
        if mtime != self._traffic_mtime:
            with open(path, "r", encoding="utf-8-sig", newline="") as handle:
                self._traffic_cache = list(csv.DictReader(handle))
            self._traffic_mtime = mtime
        return self._traffic_cache

    @staticmethod
    def _sum(rows: list[dict], key: str) -> float:
        return sum(_num(row.get(key, 0)) for row in rows)

    def _content_catalog(self) -> list[dict]:
        """Merge full content performance with the local blue-link fact table.

        The hand-off report deliberately keeps only 500 link rows for display.
        The dashboard must instead aggregate the complete local CSV so that a
        date filter and all drill-downs use the same content universe.
        """
        paths = (config.CONTENT_FACTS_PATH, config.LINK_FACTS_PATH, self.report_path)
        try:
            mtimes = tuple(os.path.getmtime(path) for path in paths)
        except OSError:
            mtimes = (-1.0, -1.0, -1.0)
        if self._content_cache and mtimes == self._content_cache_mtimes:
            return self._content_cache

        catalog: dict[str, dict] = {}
        ranked = self.get_report().get("overview_rankings", {}).get("content", {}).get("rows", []) or []
        for row in ranked:
            content_id = str(row.get("content_id") or "")
            if not content_id:
                continue
            catalog[content_id] = {
                "content_id": content_id,
                "title": row.get("title") or "[视频内容]",
                "url": row.get("content_url") or "",
                "creator_id": str(row.get("author_uid") or ""),
                "creator_name": row.get("author_name") or "",
                "date": str(row.get("published_date") or "")[:10],
                "play_count": int(_num(row.get("play_count"))),
                "interaction_count": int(_num(row.get("interaction_count"))),
                "blue_link_count": 0,
                "thunderbird_link_count": 0,
                "brands": set(),
                "brand_link_counts": defaultdict(int),
            }
        if os.path.exists(config.CONTENT_FACTS_PATH):
            with open(config.CONTENT_FACTS_PATH, "r", encoding="utf-8-sig", newline="") as handle:
                for row in csv.DictReader(handle):
                    content_id = str(row.get("content_id") or "")
                    if not content_id:
                        continue
                    entry = catalog.setdefault(content_id, {
                        "content_id": content_id, "title": "[视频内容]", "url": "", "creator_id": "",
                        "creator_name": "", "date": "", "play_count": 0, "interaction_count": 0,
                        "blue_link_count": 0, "thunderbird_link_count": 0, "brands": set(),
                        "brand_link_counts": defaultdict(int),
                    })
                    entry.update({
                        "title": row.get("title") or entry["title"],
                        "url": row.get("content_url") or entry["url"],
                        "creator_id": str(row.get("author_uid") or entry["creator_id"]),
                        "creator_name": row.get("author_name") or entry["creator_name"],
                        "date": str(row.get("published_at") or entry["date"])[:10],
                    })
        if os.path.exists(config.LINK_FACTS_PATH):
            with open(config.LINK_FACTS_PATH, "r", encoding="utf-8-sig", newline="") as handle:
                for row in csv.DictReader(handle):
                    content_id = str(row.get("content_id") or "")
                    if not content_id:
                        continue
                    entry = catalog.setdefault(content_id, {
                        "content_id": content_id, "title": "[视频内容]", "url": "", "creator_id": "",
                        "creator_name": "", "date": "", "play_count": 0, "interaction_count": 0,
                        "blue_link_count": 0, "thunderbird_link_count": 0, "brands": set(),
                        "brand_link_counts": defaultdict(int),
                    })
                    entry["blue_link_count"] += 1
                    brand = str(row.get("brand") or "").strip()
                    if brand:
                        entry["brands"].add(brand)
                        entry.setdefault("brand_link_counts", defaultdict(int))[brand] += 1
                    if brand == "雷鸟":
                        entry["thunderbird_link_count"] += 1
        self._content_cache = list(catalog.values())
        self._content_cache_mtimes = mtimes
        return self._content_cache

    def _filter_content(self, start: date, end: date, scope: str = "all", brand: Optional[str] = None,
                        creator_name: Optional[str] = None, creator_id: Optional[str] = None,
                        exact_date: Optional[str] = None) -> list[dict]:
        target_date = self._date(exact_date) if exact_date else None
        # Accept both current callers and stale cached pages that URL-encoded a
        # Chinese brand/creator before fetch encoded it a second time.
        brand = unquote(unquote(brand)).strip() if brand else None
        creator_name = unquote(unquote(creator_name)).strip() if creator_name else None
        creator_id = unquote(unquote(creator_id)).strip() if creator_id else None
        result = []
        for row in self._content_catalog():
            row_date = self._date(row.get("date"))
            if not row_date or not (start <= row_date <= end):
                continue
            if target_date and row_date != target_date:
                continue
            if scope == "thunderbird" and not row.get("thunderbird_link_count"):
                continue
            if brand and brand.casefold() not in {str(item).casefold() for item in row.get("brands", set())}:
                continue
            if creator_id and creator_id != row.get("creator_id"):
                continue
            if creator_name and creator_name not in (row.get("creator_name") or ""):
                continue
            result.append(row)
        return result

    def get_overview(self) -> dict:
        r = self.get_report()
        o = r.get("overview", {})
        pen_7d = o.get("commission_penetration_last_7_days", 0)
        pen_str = f"{pen_7d * 100:.1f}%" if isinstance(pen_7d, (int, float)) else str(pen_7d)
        completion_rate = o.get("h2_alliance_commission_completion_rate", "")
        cr_str = f"{completion_rate * 100:.1f}%" if isinstance(completion_rate, (int, float)) else str(completion_rate)
        kpis = [
            {"key": "content_count", "label": "内容数", "value": o.get("content_count", 0), "unit": "条",
             "sub_label": "雷鸟/行业", "sub_value": f"{o.get('thunderbird_linked_content_count', 0)} / {o.get('all_monitored_content_count', 0)} ({o.get('metric_coverage', '')})", "trend": ""},
            {"key": "link_count", "label": "蓝链数", "value": o.get("thunderbird_link_count", 0), "unit": "个",
             "sub_label": "雷鸟品牌达人", "sub_value": f"监播达人 {o.get('thunderbird_creator_count', 0)}人", "trend": ""},
            {"key": "store_traffic", "label": "店铺流量", "value": o.get("sales_quantity", 0), "unit": "单",
             "sub_label": "联盟佣金占比", "sub_value": f"CPS渗透率 {pen_str}", "trend": ""},
            {"key": "jd_orders", "label": "京东联盟订单", "value": o.get("sales_quantity", 0), "unit": "单",
             "sub_label": "H2目标进度", "sub_value": f"完成率 {cr_str}", "trend": ""},
            {"key": "jd_amount", "label": "京东联盟金额", "value": round(o.get("commission_amount_last_7_days", 0)), "unit": "元",
             "sub_label": "近7天总佣金", "sub_value": f"H2目标 ¥{_num(o.get('h2_alliance_commission_target', 0)):,.0f}", "trend": ""},
        ]
        return {"kpis": kpis, "date_range": o.get("period", ""), "generated_at": r.get("generated_at", ""),
                "latest_data_date": r.get("latest_data_date", "")}

    def get_top_creators(self, limit: int = 10, start_date: Optional[str] = None,
                         end_date: Optional[str] = None, scope: str = "all") -> list:
        start, end = self._resolve_period(start_date, end_date)
        grouped: dict[str, dict] = {}
        for row in self._filter_content(start, end, scope=scope):
            key = row.get("creator_id") or row.get("creator_name") or "未知达人"
            entry = grouped.setdefault(key, {
                "creator_id": row.get("creator_id", ""), "creator_name": row.get("creator_name") or "未知达人",
                "content_count": 0, "play_count": 0, "interaction_count": 0,
                "blue_link_count": 0, "thunderbird_link_count": 0,
                # Creator-level sales attribution is not date-granular in the source.
                "thunderbird_conversion": None,
            })
            entry["content_count"] += 1
            for metric in ("play_count", "interaction_count", "blue_link_count", "thunderbird_link_count"):
                entry[metric] += int(_num(row.get(metric)))
        rows = sorted(grouped.values(), key=lambda item: (item["play_count"], item["thunderbird_link_count"]), reverse=True)
        return rows[:limit]

    @staticmethod
    def _change_pct(current: float, previous: float) -> Optional[float]:
        if not previous:
            return None if current else 0.0
        return round((current - previous) / abs(previous) * 100, 2)

    def get_creator_comparison(self, start_date: Optional[str] = None,
                               end_date: Optional[str] = None, scope: str = "all") -> list:
        start, end = self._resolve_period(start_date, end_date)
        previous_start, previous_end = self._previous_period(start, end)
        metrics = ("content_count", "play_count", "interaction_count", "blue_link_count", "thunderbird_link_count")

        def aggregate(period_start: date, period_end: date) -> dict[str, dict]:
            grouped: dict[str, dict] = {}
            for row in self._filter_content(period_start, period_end, scope=scope):
                key = row.get("creator_id") or row.get("creator_name") or "未知达人"
                entry = grouped.setdefault(key, {"creator_id": row.get("creator_id", ""),
                    "creator_name": row.get("creator_name") or "未知达人", **{m: 0 for m in metrics}})
                entry["content_count"] += 1
                for metric in metrics[1:]:
                    entry[metric] += int(_num(row.get(metric)))
            return grouped

        current, previous = aggregate(start, end), aggregate(previous_start, previous_end)
        result = []
        for key in set(current) | set(previous):
            fallback = previous.get(key, {})
            cur = current.get(key) or {"creator_id": fallback.get("creator_id", ""),
                "creator_name": fallback.get("creator_name", "未知达人"), **{m: 0 for m in metrics}}
            prev = previous.get(key, {m: 0 for m in metrics})
            item = {"creator_id": cur.get("creator_id", ""), "creator_name": cur.get("creator_name", "未知达人"),
                    "thunderbird_conversion": None, "previous_thunderbird_conversion": None,
                    "conversion_change_pct": None, "conversion_data_status": "unavailable"}
            for metric in metrics:
                item[metric] = cur.get(metric, 0)
                item[f"previous_{metric}"] = prev.get(metric, 0)
                item[f"{metric}_change_pct"] = self._change_pct(cur.get(metric, 0), prev.get(metric, 0))
            result.append(item)
        return sorted(result, key=lambda item: (item["play_count"], item["thunderbird_link_count"]), reverse=True)

    def get_brand_rankings(self, start_date: Optional[str] = None, end_date: Optional[str] = None,
                           limit: int = 20) -> list:
        start, end = self._resolve_period(start_date, end_date)
        grouped: dict[str, dict] = {}
        seen: dict[str, set[str]] = defaultdict(set)
        for row in self._filter_content(start, end):
            for brand in row.get("brands", set()):
                entry = grouped.setdefault(brand, {
                    "brand_name": brand, "content_count": 0, "play_count": 0, "interaction_count": 0,
                    "link_count": 0, "thunderbird_link_count": 0, "creator_count": 0,
                })
                if row["content_id"] not in seen[brand]:
                    entry["content_count"] += 1
                    entry["play_count"] += int(_num(row.get("play_count")))
                    entry["interaction_count"] += int(_num(row.get("interaction_count")))
                    seen[brand].add(row["content_id"])
                entry["link_count"] += int(_num(row.get("brand_link_counts", {}).get(brand, 0)))
                entry["thunderbird_link_count"] += int(_num(row.get("thunderbird_link_count"))) if brand == "雷鸟" else 0
                if row.get("creator_id"):
                    seen.setdefault(f"creator:{brand}", set()).add(row["creator_id"])
        for brand, entry in grouped.items():
            entry["creator_count"] = len(seen.get(f"creator:{brand}", set()))
        return sorted(grouped.values(), key=lambda item: (item["play_count"], item["link_count"]), reverse=True)[:limit]

    def get_daily_curve(self, start_date: Optional[str] = None, end_date: Optional[str] = None) -> dict:
        start, end = self._resolve_period(start_date, end_date, days=14)
        daily = self._period_daily(start, end)
        labels, cnt, lk, sal, amt = [], [], [], [], []
        for d in daily:
            dt = d.get("date", "")
            labels.append(dt[5:])
            cnt.append(int(_num(d.get("new_content", 0))))
            lk.append(int(_num(d.get("thunderbird_link_count", 0))))
            sal.append(int(_num(d.get("sales_quantity", 0))))
            amt.append(round(_num(d.get("commission_amount", 0))))
        return {"labels": labels, "content_count": cnt, "link_count": lk, "sales_count": sal, "amount": amt}

    def get_action_plan(self, start_date: Optional[str] = None, end_date: Optional[str] = None) -> dict:
        r = self.get_report()
        ap = r.get("action_plan", {}) or {}
        kv = {}
        for k, v in ap.items():
            n = len(k)
            if n == 3:
                kv["pending"] = v
            elif n == 4:
                kv["churn"] = v
            elif n == 6:
                if "highval" not in kv:
                    kv["highval"] = v
                else:
                    kv["policy"] = v
            else:
                kv[f"other_{n}"] = v
        plan = {"待建联": kv.get("pending", []), "流失预警": kv.get("churn", r.get("creator_churn", [])),
                "高价值基本盘": kv.get("highval", []), "政策倾斜建议": kv.get("policy", [])}
        selected_start, selected_end = self._resolve_period(start_date, end_date)
        selected_comparisons = self.get_creator_comparison(selected_start.isoformat(), selected_end.isoformat())
        creators_by_id = {str(row.get("creator_id") or ""): row for row in selected_comparisons if row.get("creator_id")}
        creators_by_name = {row["creator_name"]: row for row in selected_comparisons}
        library = self.get_report().get("creator_library", []) or []
        library_by_id = {str(row.get("author_uid") or ""): row for row in library if row.get("author_uid")}
        library_by_name = {str(row.get("author_name") or "").strip(): row for row in library if row.get("author_name")}
        pending_enriched = []
        for row in plan["待建联"]:
            uid = str(row.get("creator_uid") or row.get("creator_id") or "")
            name = row.get("creator_name") or row.get("creator_or_pin") or row.get("name") or "未知达人"
            identity = library_by_id.get(uid) or library_by_name.get(name) or {}
            # A promoter PIN means the creator is already linked. The outreach
            # queue is exclusively for creators without any PIN mapping.
            if identity.get("promoter_pins"):
                continue
            metric = creators_by_id.get(uid) or creators_by_name.get(name) or {}
            evidence = row.get("evidence") or {}
            content_count = int(_num(metric.get("content_count", evidence.get("content_count", 0))))
            play_count = int(_num(metric.get("play_count", evidence.get("play_count", 0))))
            interaction_count = int(_num(metric.get("interaction_count", evidence.get("interaction_count", 0))))
            blue_link_count = int(_num(metric.get("blue_link_count", evidence.get("blue_link_count", 0))))
            thunderbird_link_count = int(_num(metric.get("thunderbird_link_count", evidence.get("thunderbird_link_count", 0))))
            reasons = []
            if play_count >= 1_000_000:
                reasons.append(f"周期播放{play_count / 10000:.1f}万，属于高流量达人")
            elif play_count:
                reasons.append(f"周期播放{play_count / 10000:.1f}万，具备内容触达能力")
            if content_count >= 10:
                reasons.append(f"产出稳定（{content_count}条内容）")
            if blue_link_count and not thunderbird_link_count:
                reasons.append(f"已有{blue_link_count}条竞品/行业蓝链，但暂无雷鸟蓝链")
            elif thunderbird_link_count:
                reasons.append(f"已有{thunderbird_link_count}条雷鸟蓝链，可进一步深化合作")
            if not reasons:
                reasons.append(row.get("reason") or "行业相关达人，建议补充建联并持续观察")
            pending_enriched.append({**row, **metric, "creator_uid": uid, "creator_name": name,
                "content_count": content_count, "play_count": play_count,
                "interaction_count": interaction_count, "blue_link_count": blue_link_count,
                "thunderbird_link_count": thunderbird_link_count,
                "recommendation_reason": "；".join(reasons[:3])})
        plan["待建联"] = sorted(pending_enriched, key=lambda item: (item["play_count"], item["content_count"]), reverse=True)
        high_value_enriched = []
        for row in plan["高价值基本盘"]:
            uid = str(row.get("creator_uid") or row.get("creator_id") or "")
            name = row.get("creator_name") or row.get("creator_or_pin") or row.get("name") or "未知达人"
            metric = creators_by_id.get(uid) or creators_by_name.get(name) or {}
            evidence = row.get("evidence") or {}
            values = {key: int(_num(metric.get(key, evidence.get(key, 0)))) for key in
                ("content_count", "play_count", "interaction_count", "blue_link_count", "thunderbird_link_count")}
            reasons = []
            if values["thunderbird_link_count"]:
                reasons.append(f'周期内有{values["thunderbird_link_count"]}条雷鸟蓝链，合作基础稳定')
            if values["play_count"]:
                reasons.append(f'周期播放{values["play_count"] / 10000:.1f}万')
            if values["content_count"]:
                reasons.append(f'产出{values["content_count"]}条内容')
            if not reasons:
                reasons.append(row.get("reason") or "历史雷鸟蓝链贡献靠前，已建立基本盘")
            high_value_enriched.append({**row, **metric, **values, "creator_uid": uid,
                "creator_name": name, "recommendation_reason": "；".join(reasons)})
        plan["高价值基本盘"] = high_value_enriched
        sales_changes = {}
        for row in self.get_report().get("creator_churn", []) or []:
            name = str(row.get("creator_or_pin") or "").strip()
            current_sales, previous_sales = _num(row.get("current_sales")), _num(row.get("previous_sales"))
            sales_changes[name] = {"current_sales": int(current_sales), "previous_sales": int(previous_sales),
                "sales_change_pct": self._change_pct(current_sales, previous_sales)}

        churn, positive = [], []
        for metric in selected_comparisons:
            name = metric.get("creator_name") or "未知达人"
            identity = library_by_id.get(str(metric.get("creator_id") or "")) or library_by_name.get(name) or {}
            if not identity.get("promoter_pins"):
                continue
            sales = sales_changes.get(name, {})
            changes = {"播放": metric.get("play_count_change_pct"), "蓝链": metric.get("blue_link_count_change_pct"),
                       "销量": sales.get("sales_change_pct")}
            declines = [label for label, value in changes.items() if value is not None and value < 0]
            rises = [label for label, value in changes.items() if value is not None and value > 0]
            item = {**metric, **sales, "creator_name": name, "metric_data_status": "ready",
                    "status": "、".join(f"{label}下降" for label in declines) if declines else "表现增长",
                    "recommended_action": "针对下降指标复盘内容、挂链和销售承接" if declines else "持续维护并放大增长指标"}
            # Growth takes precedence for mixed movements: a creator with a
            # rising play/link/sales signal belongs in the value base rather
            # than appearing simultaneously as churn risk.
            if rises:
                reasons = [f'{label}同比上升{changes[label]:.1f}%' for label in rises]
                positive.append({**item, "recommendation_reason": "；".join(reasons)})
            elif declines:
                churn.append(item)
        plan["流失预警"] = sorted(churn, key=lambda item: min(
            [value for value in (item.get("play_count_change_pct"), item.get("blue_link_count_change_pct"), item.get("sales_change_pct")) if value is not None] or [0]))
        churn_ids = {str(row.get("creator_id") or "") for row in churn if row.get("creator_id")}
        churn_names = {row.get("creator_name") for row in churn}
        plan["高价值基本盘"] = [row for row in plan["高价值基本盘"]
            if str(row.get("creator_uid") or row.get("creator_id") or "") not in churn_ids
            and row.get("creator_name") not in churn_names]
        positive_by_id = {str(row.get("creator_id") or ""): row for row in positive if row.get("creator_id")}
        positive_by_name = {row.get("creator_name"): row for row in positive}
        plan["高价值基本盘"] = [{**row, **(positive_by_id.get(str(row.get("creator_uid") or row.get("creator_id") or ""))
            or positive_by_name.get(row.get("creator_name")) or {})} for row in plan["高价值基本盘"]]
        existing_ids = {str(row.get("creator_uid") or row.get("creator_id") or "") for row in plan["高价值基本盘"]}
        existing_names = {row.get("creator_name") for row in plan["高价值基本盘"]}
        plan["高价值基本盘"].extend(row for row in positive
            if str(row.get("creator_id") or "") not in existing_ids and row.get("creator_name") not in existing_names)
        plan["counts"] = {k: len(v) for k, v in plan.items()}
        return plan

    def get_today_tasks(self) -> dict:
        p = self.get_action_plan()
        return {"items": [
            {"title": "达人建联", "desc": "优先联系头部竞品强关联达人（待建联列表前10名）", "status": "pending", "priority": "high", "count": p["counts"].get("待建联", 0)},
            {"title": "达人运营", "desc": "跟进流失预警达人（上月销售下滑需定向触达）", "status": "pending", "priority": "high", "count": p["counts"].get("流失预警", 0)},
            {"title": "政策建议", "desc": "评估高价值达人的政策倾斜方案（ROI台账待完善）", "status": "pending", "priority": "medium", "count": p["counts"].get("高价值基本盘", 0)},
        ]}

    def _daily_periods(self, days: int = 7, start_date: Optional[str] = None,
                       end_date: Optional[str] = None) -> tuple[list[dict], list[dict]]:
        start, end = self._resolve_period(start_date, end_date, days=days)
        previous_start, previous_end = self._previous_period(start, end)
        return self._period_daily(start, end), self._period_daily(previous_start, previous_end)

    def get_kpi_cards(self, days: int = 7, start_date: Optional[str] = None,
                      end_date: Optional[str] = None) -> dict:
        start, end = self._resolve_period(start_date, end_date, days=days)
        current, previous = self._daily_periods(days, start.isoformat(), end.isoformat())
        cur_sum = lambda key: sum(_num(d.get(key, 0)) for d in current)
        prev_sum = lambda key: sum(_num(d.get(key, 0)) for d in previous)

        def ratio(cur, tot):
            return (cur / tot * 100) if tot else 0.0

        def yoy(cur_r, prev_r):
            if prev_r:
                chg = (cur_r - prev_r) / abs(prev_r) * 100
            else:
                chg = cur_r if cur_r else 0.0
            return round(chg, 2), ("↑" if chg >= 0 else "↓")

        specs = []

        # Every numerator and denominator is derived from the selected period.
        # The previous implementation mixed a report-wide overview total with
        # a period-limited previous value, producing a misleading comparison.
        c_cur = cur_sum("thunderbird_linked_content_count")
        c_tot = cur_sum("all_monitored_content_count")
        p_cur = prev_sum("thunderbird_linked_content_count")
        p_tot = prev_sum("all_monitored_content_count")
        chg, direction = yoy(ratio(c_cur, c_tot), ratio(p_cur, p_tot))
        specs.append({"key": "content_count", "label": "内容数", "current_value": round(c_cur), "total_value": round(c_tot),
                      "ratio": round(ratio(c_cur, c_tot), 2), "prev_period_value": round(p_cur),
                      "prev_ratio": round(ratio(p_cur, p_tot), 2), "yoy_change_pct": chg, "yoy_direction": direction,
                      "unit": "条", "click_count": round(c_cur), "detail_scope": "thunderbird", "data_status": "ready"})

        c_cur = cur_sum("thunderbird_link_count")
        c_tot = cur_sum("total_blue_link_count")
        p_cur = prev_sum("thunderbird_link_count")
        p_tot = prev_sum("total_blue_link_count")
        chg, direction = yoy(ratio(c_cur, c_tot), ratio(p_cur, p_tot))
        specs.append({"key": "link_count", "label": "蓝链数", "current_value": round(c_cur), "total_value": round(c_tot),
                      "ratio": round(ratio(c_cur, c_tot), 2), "prev_period_value": round(p_cur),
                      "prev_ratio": round(ratio(p_cur, p_tot), 2), "yoy_change_pct": chg, "yoy_direction": direction,
                      "unit": "个", "click_count": round(c_cur), "detail_scope": "thunderbird", "data_status": "ready"})

        # Keep the operational overview usable with the fields that
        # are currently available.  Each card declares its exact source metric
        # so consumers do not mistake these values for unavailable GMV/UV data.
        for key, label, numerator_key, denominator_key, unit, metric_note in [
            ("play_count", "播放量", "play_count", "all_monitored_play_count", "次", "雷鸟相关播放/行业监测播放"),
            ("jd_orders", "京东联盟订单", "sales_quantity", "jdsz_transaction_item_quantity", "单", "联盟有效销量/京准通成交商品件数"),
            ("jd_amount", "京东联盟佣金", "commission_amount", "jdsz_transaction_amount", "元", "联盟佣金/京准通成交金额"),
        ]:
            c_cur, c_tot = cur_sum(numerator_key), cur_sum(denominator_key)
            p_cur, p_tot = prev_sum(numerator_key), prev_sum(denominator_key)
            chg, direction = yoy(ratio(c_cur, c_tot), ratio(p_cur, p_tot))
            specs.append({"key": key, "label": label, "current_value": round(c_cur, 2),
                          "total_value": round(c_tot, 2), "ratio": round(ratio(c_cur, c_tot), 2),
                          "prev_period_value": round(p_cur, 2), "prev_ratio": round(ratio(p_cur, p_tot), 2),
                          "yoy_change_pct": chg, "yoy_direction": direction, "unit": unit,
                          "click_count": round(c_cur), "detail_scope": "all", "data_status": "ready",
                          "metric_note": metric_note})

        # User-defined key-source traffic: search visitors plus off-site
        # visitors, divided by all-site visitors. Search is a subset of indoor
        # traffic, so this is intentionally presented as a focus-source share,
        # not as JD's original total-traffic composition.
        c_cur = cur_sum("search_visitors") + cur_sum("outdoor_visitors")
        c_tot = cur_sum("total_visitors")
        p_cur = prev_sum("search_visitors") + prev_sum("outdoor_visitors")
        p_tot = prev_sum("total_visitors")
        chg, direction = yoy(ratio(c_cur, c_tot), ratio(p_cur, p_tot))
        specs.insert(3, {"key": "store_traffic", "label": "店铺流量", "current_value": round(c_cur, 2),
                         "total_value": round(c_tot, 2), "ratio": round(ratio(c_cur, c_tot), 2),
                         "prev_period_value": round(p_cur, 2), "prev_ratio": round(ratio(p_cur, p_tot), 2),
                         "yoy_change_pct": chg, "yoy_direction": direction, "unit": "人",
                         "click_count": round(c_cur), "detail_scope": "all", "data_status": "ready",
                         "metric_note": "搜索流量+站外流量/全站流量", "accent": "traffic"})

        period_start, period_end = self._period_label(start, end)
        return {"cards": specs, "period_start": period_start, "period_end": period_end,
                "previous_period_start": self._previous_period(start, end)[0].isoformat(),
                "previous_period_end": self._previous_period(start, end)[1].isoformat()}

    def _top_content_lookup(self) -> dict:
        lookup = {}
        dct = self.get_report().get("daily_content_top", {}) or {}
        for day_rows in dct.values():
            if not isinstance(day_rows, list):
                continue
            for row in day_rows:
                cid = row.get("content_id", "")
                if cid:
                    lookup[cid] = row
        return lookup

    def _enrich_content_row(self, row: dict, top: dict) -> dict:
        cid = row.get("content_id", "")
        meta = top.get(cid, {}) or {}
        return {
            "title": row.get("title") or meta.get("title") or "[视频内容]",
            "url": row.get("content_url") or meta.get("content_url") or "",
            "creator_name": row.get("author_name", "") or meta.get("author_name", "") or "",
            "play_count": int(_num(row.get("play_count", 0)) or _num(meta.get("play_count", 0))),
            "interaction_count": 0,
            "blue_link_count": int(_num(row.get("blue_link_count", 0))),
            "thunderbird_link_count": int(_num(row.get("thunderbird_link_count", 0))),
            "date": str(row.get("published_at", ""))[:10],
        }

    def get_content_detail(self, brand: Optional[str] = None, creator_name: Optional[str] = None,
                           creator_id: Optional[str] = None, start_date: Optional[str] = None,
                           end_date: Optional[str] = None, exact_date: Optional[str] = None,
                           scope: str = "all", limit: int = 50) -> list:
        start, end = self._resolve_period(start_date, end_date)
        rows = self._filter_content(start, end, scope=scope, brand=brand, creator_name=creator_name,
                                    creator_id=creator_id, exact_date=exact_date)
        rows.sort(key=lambda row: row["play_count"], reverse=True)
        return [{key: (sorted(value) if key == "brands" else value) for key, value in row.items()
                 if key != "brand_link_counts"}
                for row in rows[:limit]]

    def get_creator_detail(self, creator_name: str, start_date: Optional[str] = None,
                           end_date: Optional[str] = None) -> list:
        # Keep the legacy route name but now return the requested creator's
        # actual content records, not a misleading daily summary.
        return self.get_content_detail(creator_name=creator_name, start_date=start_date,
                                       end_date=end_date, limit=500)

    def get_daily_with_period(self, days: int = 7, start_date: Optional[str] = None,
                              end_date: Optional[str] = None) -> dict:
        current, previous = self._daily_periods(days, start_date, end_date)
        keys = ["date", "new_content", "thunderbird_linked_content_count", "all_monitored_content_count",
                "thunderbird_link_count", "total_blue_link_count", "sales_quantity", "commission_amount",
                "play_count", "all_monitored_play_count", "jdsz_transaction_item_quantity",
                "jdsz_transaction_amount", "search_visitors", "outdoor_visitors", "total_visitors"]

        def clean(items):
            return [{k: d.get(k, 0) for k in keys} for d in items]

        return {"current_period": clean(current), "previous_period": clean(previous)}

    @staticmethod
    def _correlation(rows: list[dict], left: str, right: str) -> Optional[float]:
        pairs = [(_num(row.get(left)), _num(row.get(right))) for row in rows]
        if len(pairs) < 3:
            return None
        xs, ys = zip(*pairs)
        mx, my = sum(xs) / len(xs), sum(ys) / len(ys)
        numerator = sum((x - mx) * (y - my) for x, y in pairs)
        dx = math.sqrt(sum((x - mx) ** 2 for x in xs))
        dy = math.sqrt(sum((y - my) ** 2 for y in ys))
        return round(numerator / (dx * dy), 3) if dx and dy else None

    def get_auto_analysis(self, start_date: Optional[str] = None,
                          end_date: Optional[str] = None, days: int = 7) -> dict:
        start, end = self._resolve_period(start_date, end_date, days)
        current, previous = self._daily_periods(days, start.isoformat(), end.isoformat())
        metric_keys = {"content": "new_content", "play": "play_count", "blue_link": "total_blue_link_count",
                       "traffic": "total_visitors", "orders": "sales_quantity", "commission": "commission_amount"}
        totals = {name: self._sum(current, key) for name, key in metric_keys.items()}
        previous_totals = {name: self._sum(previous, key) for name, key in metric_keys.items()}
        changes = {name: (round((value - previous_totals[name]) / abs(previous_totals[name]) * 100, 2)
                          if previous_totals[name] else None) for name, value in totals.items()}
        correlation_rows = [{**row, "store_traffic": _num(row.get("search_visitors")) + _num(row.get("outdoor_visitors"))}
                            for row in current]
        pairs = [("内容数与店铺流量", "new_content", "store_traffic", "店铺流量"),
                 ("蓝链数与店铺流量", "total_blue_link_count", "store_traffic", "店铺流量"),
                 ("播放量与店铺流量", "play_count", "store_traffic", "店铺流量"),
                 ("内容数与京东联盟订单", "new_content", "sales_quantity", "京东联盟订单"),
                 ("蓝链数与京东联盟订单", "total_blue_link_count", "sales_quantity", "京东联盟订单"),
                 ("播放量与京东联盟订单", "play_count", "sales_quantity", "京东联盟订单")]
        correlations = [{"label": label, "value": self._correlation(correlation_rows, left, right), "group": group}
                        for label, left, right, group in pairs]
        valid = sorted((item for item in correlations if item["value"] is not None),
                       key=lambda item: abs(item["value"]), reverse=True)
        findings = []
        if valid:
            strongest = valid[0]
            findings.append(f'{strongest["label"]}呈{"正相关" if strongest["value"] > 0 else "负相关"}（相关系数 {strongest["value"]}）')
        for key, label in (("play", "播放量"), ("blue_link", "蓝链数"), ("orders", "联盟订单"), ("commission", "联盟佣金")):
            change = changes[key]
            if change is not None:
                findings.append(f'{label}较上一等长周期{"增长" if change >= 0 else "下降"}{abs(change):.1f}%')
        recommendations = []
        if changes.get("play") is not None and changes["play"] > 0 and (changes.get("orders") or 0) <= 0:
            recommendations.append("播放增长未同步带动订单，建议复盘高播放内容的蓝链覆盖、商品承接页与转化路径。")
        if changes.get("blue_link") is not None and changes["blue_link"] < 0:
            recommendations.append("蓝链数量下降，建议优先补齐高播放内容的蓝链并跟进高价值达人。")
        if not recommendations:
            recommendations.append("优先放大与订单正相关度最高的指标，并持续观察至少7个有效数据日。")
        return {"period_start": start.isoformat(), "period_end": end.isoformat(), "sample_days": len(current),
                "totals": totals, "changes": changes, "correlations": correlations,
                "findings": findings[:5], "recommendations": recommendations[:3],
                "quality_note": "相关性按所选周期逐日计算；店铺流量=搜索流量+站外流量。相关性不等同于因果结论，少于3个有效数据日时不计算。"}

    @staticmethod
    def _content_direction(titles: list[str]) -> str:
        themes = {
            "选购测评": ("测评", "评测", "对比", "怎么选", "选购"),
            "产品体验": ("体验", "开箱", "实测", "使用"),
            "电竞游戏": ("电竞", "游戏", "高刷", "帧率"),
            "办公生产力": ("办公", "设计", "剪辑", "生产力"),
            "价格促销": ("价格", "优惠", "值得买", "性价比", "促销"),
        }
        scores = {name: sum(any(word in str(title) for word in words) for title in titles)
                  for name, words in themes.items()}
        ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
        selected = [name for name, count in ranked if count > 0][:2]
        return "、".join(selected) if selected else "综合产品介绍"

    def get_industry_analysis(self, start_date: Optional[str] = None,
                              end_date: Optional[str] = None, days: int = 7) -> dict:
        start, end = self._resolve_period(start_date, end_date, days)
        rows = self._filter_content(start, end)
        brands = self.get_brand_rankings(start.isoformat(), end.isoformat(), 10)
        brand_details = []
        for brand in brands:
            content = max(1, int(brand["content_count"]))
            creators = max(1, int(brand["creator_count"]))
            avg_play = round(brand["play_count"] / content)
            density_score = content / max(1, max((x["content_count"] for x in brands), default=1))
            coverage_score = creators / max(1, max((x["creator_count"] for x in brands), default=1))
            hit_score = avg_play / max(1, max((x["play_count"] / max(1, x["content_count"]) for x in brands), default=1))
            driver = max(((density_score, "内容密度"), (coverage_score, "达人覆盖"), (hit_score, "单条爆款")))[1]
            brand_details.append({**brand, "avg_play": avg_play, "growth_driver": driver})

        creators: dict[str, dict] = {}
        for row in rows:
            key = row.get("creator_id") or row.get("creator_name") or "未知达人"
            item = creators.setdefault(key, {"creator_id": row.get("creator_id", ""),
                "creator_name": row.get("creator_name") or "未知达人", "play_count": 0,
                "content_count": 0, "thunderbird_link_count": 0, "titles": []})
            item["play_count"] += int(_num(row.get("play_count")))
            item["content_count"] += 1
            item["thunderbird_link_count"] += int(_num(row.get("thunderbird_link_count")))
            item["titles"].append(str(row.get("title") or ""))
        top_creators = sorted(creators.values(), key=lambda item: item["play_count"], reverse=True)[:10]
        for item in top_creators:
            item["content_direction"] = self._content_direction(item.pop("titles"))
            item["cooperation_status"] = "已建联合作" if item["thunderbird_link_count"] > 0 else "待建联"

        own_rows = [row for row in rows if "雷鸟" in row.get("brands", set())]
        own_direction = self._content_direction([str(row.get("title") or "") for row in own_rows])
        own = {"content_count": len(own_rows), "play_count": sum(int(_num(row.get("play_count"))) for row in own_rows),
               "interaction_count": sum(int(_num(row.get("interaction_count"))) for row in own_rows),
               "creator_count": len({row.get("creator_id") or row.get("creator_name") for row in own_rows}),
               "content_direction": own_direction}
        pending_top = [item for item in top_creators if item["cooperation_status"] == "待建联"]
        recommendations = []
        if pending_top:
            recommendations.append(f'优先建联行业Top达人“{pending_top[0]["creator_name"]}”，其主要方向为{pending_top[0]["content_direction"]}。')
        if brand_details:
            leader = brand_details[0]
            recommendations.append(f'{leader["brand_name"]}当前主要由{leader["growth_driver"]}驱动，雷鸟应针对该路径配置内容与达人资源。')
        recommendations.append(f'本品当前内容方向为{own_direction}，建议围绕高播放方向增加内容密度并同步补齐蓝链。')
        return {"module": "industry", "title": "AI 行业竞争与CPS策略分析",
                "period_start": start.isoformat(), "period_end": end.isoformat(), "sample_days": len({row.get("date") for row in rows}),
                "brand_drivers": brand_details[:6], "top_creators": top_creators, "own_content": own,
                "recommendations": recommendations, "quality_note": "驱动类型根据内容量、达人覆盖及单条平均播放的相对强度判断。"}

    def get_creator_workbench_analysis(self, start_date: Optional[str] = None,
                                       end_date: Optional[str] = None, days: int = 7) -> dict:
        start, end = self._resolve_period(start_date, end_date, days)
        plan = self.get_action_plan(start.isoformat(), end.isoformat())
        counts = plan.get("counts", {})
        pending = plan.get("待建联", [])
        churn = plan.get("流失预警", [])
        high_value = plan.get("高价值基本盘", [])
        actions = []
        if pending:
            actions.append(f'优先处理前{min(10, len(pending))}位待建联达人，先生成个性化话术并人工确认后发送。')
        if churn:
            actions.append(f'对{len(churn)}位流失预警达人按播放与蓝链降幅排序回访。')
        if high_value:
            actions.append(f'对{len(high_value)}位高价值达人建立月度维护与政策倾斜清单。')
        return {"module": "creator", "title": "AI 达人运营分析", "period_start": start.isoformat(),
                "period_end": end.isoformat(), "sample_days": (end - start).days + 1,
                "counts": counts, "actions": actions, "recommendations": actions,
                "quality_note": "建联动作需在发送前确认达人UID、话术和发送频率。"}

    def get_ai_context(self) -> dict:
        return {"overview": self.get_overview(), "top_creators": self.get_top_creators(10),
                "brand_rankings": self.get_brand_rankings(), "creators": self.get_report().get("creators", [])[:50],
                "action_plan": self.get_action_plan(), "daily_curve": self.get_daily_curve(),
                "kpi_cards": self.get_kpi_cards(7), "daily_with_period": self.get_daily_with_period(7)}


loader = DataLoader()


async def call_deepseek(system_prompt: str, user_prompt: str) -> str:
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            f"{config.DEEPSEEK_BASE_URL}/chat/completions",
            headers={"Authorization": f"Bearer {config.DEEPSEEK_API_KEY}", "Content-Type": "application/json"},
            json={"model": config.DEEPSEEK_MODEL,
                  "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                  "temperature": 0.7, "max_tokens": 800})
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]


@app.get("/")
async def serve_dashboard():
    return FileResponse(os.path.join(config.PROJECT_DIR, "dashboard.html"), media_type="text/html; charset=utf-8")


@app.get("/echarts.min.js")
async def serve_echarts():
    return FileResponse(os.path.join(config.PROJECT_DIR, "echarts.min.js"), media_type="application/javascript")


@app.get("/favicon.ico")
async def serve_favicon():
    return JSONResponse(status_code=200, content={})


@app.get("/api/overview")
def api_overview():
    return loader.get_overview()


@app.get("/api/top-creators")
def api_top_creators(limit: int = Query(10, ge=1, le=5000), start_date: Optional[str] = None,
                     end_date: Optional[str] = None, scope: str = "all"):
    return loader.get_top_creators(limit, start_date, end_date, scope)


@app.get("/api/brand-rankings")
def api_brand_rankings(start_date: Optional[str] = None, end_date: Optional[str] = None,
                       limit: int = Query(20, ge=1, le=100)):
    return loader.get_brand_rankings(start_date, end_date, limit)


@app.get("/api/action-plan")
def api_action_plan(start_date: Optional[str] = None, end_date: Optional[str] = None):
    return loader.get_action_plan(start_date, end_date)

@app.get("/api/creator-comparison")
def api_creator_comparison(start_date: Optional[str] = None, end_date: Optional[str] = None,
                           scope: str = "all"):
    return loader.get_creator_comparison(start_date, end_date, scope)


@app.get("/api/daily-curve")
def api_daily_curve(start_date: Optional[str] = None, end_date: Optional[str] = None):
    return loader.get_daily_curve(start_date, end_date)


@app.get("/api/today-tasks")
def api_today_tasks():
    return loader.get_today_tasks()


@app.get("/api/kpi-cards")
def api_kpi_cards(period_select: int = Query(7, ge=1, le=60), start_date: Optional[str] = None,
                  end_date: Optional[str] = None):
    return loader.get_kpi_cards(period_select, start_date, end_date)


@app.get("/api/content-detail")
def api_content_detail(brand: Optional[str] = None, creator_name: Optional[str] = None,
                       creator_id: Optional[str] = None, start_date: Optional[str] = None,
                       end_date: Optional[str] = None, date: Optional[str] = None, scope: str = "all",
                       limit: int = Query(50, ge=1, le=5000)):
    return loader.get_content_detail(brand=brand, creator_name=creator_name, creator_id=creator_id,
                                     start_date=start_date, end_date=end_date, exact_date=date,
                                     scope=scope, limit=limit)


@app.get("/api/creator-detail")
def api_creator_detail(creator_name: str = Query(..., description="达人名称"),
                       start_date: Optional[str] = None, end_date: Optional[str] = None):
    return loader.get_creator_detail(creator_name, start_date, end_date)


@app.get("/api/daily-with-period")
def api_daily_with_period(days: int = Query(7, ge=1, le=60), start_date: Optional[str] = None,
                          end_date: Optional[str] = None):
    return loader.get_daily_with_period(days, start_date, end_date)


@app.get("/api/auto-analysis")
def api_auto_analysis(days: int = Query(7, ge=1, le=60), start_date: Optional[str] = None,
                      end_date: Optional[str] = None, module: str = "platform"):
    if module == "industry":
        return loader.get_industry_analysis(start_date, end_date, days)
    if module == "creator":
        return loader.get_creator_workbench_analysis(start_date, end_date, days)
    result = loader.get_auto_analysis(start_date, end_date, days)
    result.update({"module": "platform", "title": "AI 自动经营分析"})
    return result


@app.get("/api/roi-analysis")
def api_roi_analysis(tracking_days: int = Query(30, ge=1, le=365)):
    return loader.get_roi_analysis(tracking_days)


@app.post("/api/roi-import")
async def api_roi_import(file: UploadFile = File(...)):
    if not (file.filename or "").lower().endswith(".xlsx"):
        return JSONResponse({"error": "仅支持.xlsx文件"}, status_code=400)
    workbook = load_workbook(file.file, read_only=True, data_only=True)
    sheet = workbook.active
    # Some exported workbooks carry an incorrect A1-only worksheet dimension
    # although their table/data extends across multiple columns and rows.
    sheet.reset_dimensions()
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    required = ["投入达人", "投放日期", "投入金额", "推客pin"]
    if any(name not in headers for name in required):
        return JSONResponse({"error": f"表头必须包含：{'、'.join(required)}"}, status_code=400)
    positions = {name: headers.index(name) for name in required}
    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[positions["投入达人"]] or "").strip()
        if not name: continue
        raw_date = row[positions["投放日期"]]
        parsed = loader._investment_date(raw_date)
        amount = row[positions["投入金额"]]
        pins = [part.strip() for part in str(row[positions["推客pin"]] or "").splitlines() if part.strip()]
        items.append({"creator_name": name, "investment_date": parsed.isoformat() if parsed else str(raw_date or ""),
            "investment_amount": float(amount) if isinstance(amount, (int, float)) else None, "promoter_pins": pins})
    os.makedirs(os.path.dirname(config.ROI_DATA_PATH), exist_ok=True)
    with open(config.ROI_DATA_PATH, "w", encoding="utf-8") as handle:
        json.dump({"source_file": file.filename, "imported_at": datetime.now().isoformat(), "items": items}, handle, ensure_ascii=False, indent=2)
    return {"success": True, "count": len(items), "analysis": loader.get_roi_analysis(30)}


@app.post("/api/ai-generate")
async def ai_generate(data: dict):
    rt = data.get("type", "")
    ctx = data.get("context", {}) or {}
    pt = config.ANALYSIS_PROMPTS.get(rt, "")
    if not pt:
        return {"error": f"Unknown type: {rt}"}

    # Extract period — prefer explicit period_select, then deduce from context
    days = 7
    period_label = ""
    ps = data.get("period_select")
    if isinstance(ps, str) and ps.isdigit():
        days = int(ps)
    elif isinstance(ps, int) and ps > 0:
        days = ps
    else:
        ctx_ps = ctx.get("period_select")
        if isinstance(ctx_ps, str) and ctx_ps.isdigit():
            days = int(ctx_ps)
        elif isinstance(ctx_ps, int) and ctx_ps > 0:
            days = ctx_ps
        # Try to deduce from context.period label
        ctx_period = ctx.get("period", "")
        for label_val, key in [("近30天", 30), ("近14天", 14), ("近7天", 7), ("今天", 1), ("昨天", 1), ("近30天", 30)]:
            if label_val in ctx_period:
                days = key
                break

    parts = []
    
    # KPI cards (always fetched from server with correct period)
    cards = loader.get_kpi_cards(days)
    if cards.get("cards"):
        pstart = cards.get("period_start", "未知")
        pend = cards.get("period_end", "未知")
        parts.append(f"\n## KPI数据（{pstart} ~ {pend}，当前{days}天 vs 前{days}天同期对比）:")
        for c in cards["cards"]:
            yoy_str = f"{c['yoy_direction']} {abs(c['yoy_change_pct'])}% ({'增' if c['yoy_direction'] == '↑' else '减'})"
            parts.append(f"- {c['label']}: 当前{c['current_value']}，行业{c['total_value']}，占比{c['ratio']}%；同比{yoy_str}")

    # Daily curve data (current + previous period)
    daily_data = loader.get_daily_with_period(days)
    cp = daily_data.get("current_period", [])
    pp = daily_data.get("previous_period", [])
    
    if cp:
        parts.append("\n## 日度趋势:")
        for d in cp[-10:]:  # last 10 days
            parts.append(f"- {d['date']}: 内容{d['new_content']}条，蓝链{d['total_blue_link_count']}个，雷鸟蓝链{d['thunderbird_link_count']}个，销量{d['sales_quantity']}单，佣金¥{_num(d.get('commission_amount', 0)):.0f}")
        
        if len(cp) == len(pp) and pp:
            total_cp_content = sum(_num(d.get("new_content", 0)) for d in cp)
            total_pp_content = sum(_num(d.get("new_content", 0)) for d in pp)
            total_cp_sales = sum(_num(d.get("sales_quantity", 0)) for d in cp)
            total_pp_sales = sum(_num(d.get("sales_quantity", 0)) for d in pp)
            total_cp_amt = sum(_num(d.get("commission_amount", 0)) for d in cp)
            total_pp_amt = sum(_num(d.get("commission_amount", 0)) for d in pp)
            diff_c = total_cp_content - total_pp_content
            diff_s = total_cp_sales - total_pp_sales
            diff_a = total_cp_amt - total_pp_amt
            parts.append(f"- 与前期比: 内容 {'+' if diff_c >= 0 else ''}{diff_c}条，销量 {'+' if diff_s >= 0 else ''}{diff_s}单，佣金 {'+' if diff_a >= 0 else ''}{diff_a:.0f}元")

    # Top creators from context
    creators = ctx.get("top_creators", [])[:10]
    if creators:
        parts.append("\n## Top达人:")
        for cr in creators:
            parts.append(f"- {cr.get('creator_name', '')}: 雷鸟销售{cr.get('thunderbird_sales', 0)}单，佣金¥{_num(cr.get('commission', 0)):.0f}")

    # Action plan summary
    ap = ctx.get("action_plan_summary", ctx.get("action_plan", {})) or {}
    counts = ap.get("counts", {})
    if counts:
        parts.append("\n## 行动计划:")
        for k, v in counts.items():
            parts.append(f"- {k}: {v}位")
    
    # Daily curve raw data from frontend context
    dc = ctx.get("daily_curve", [])
    if isinstance(dc, list) and dc:
        parts.append(f"\n## 前端传入的日度数据（共{len(dc)}天）:")
        for item in dc[:5]:
            if isinstance(item, dict):
                parts.append(f"- {item.get('date', '')}: 内容{item.get('new_content', 0)}, 蓝链{item.get('total_blue_link_count', 0)}, 雷鸟{item.get('thunderbird_link_count', 0)}")

    try:
        user_prompt = "\n".join(parts)
        result = await call_deepseek(pt, user_prompt or "请基于现有数据分析。")
        return {"text": result.strip(), "timestamp": datetime.now().isoformat()}
    except Exception as e:
        return {"error": str(e)}


@app.on_event("startup")
async def startup():
    print("=== CPS Dashboard ===")
    print(f"   http://{config.HOST}:{config.PORT}")
    print(f"   AI: {config.DEEPSEEK_MODEL}")
