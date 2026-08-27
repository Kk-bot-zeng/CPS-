import os, sys
from openpyxl import load_workbook
import psycopg2

book = load_workbook(sys.argv[1], read_only=True, data_only=True)
sheet = book.active
rows = [row for row in sheet.iter_rows(min_row=2, values_only=True) if any(value == "#N/A" for value in row)]
conn = psycopg2.connect(os.environ["DATABASE_URL"])
with conn, conn.cursor() as cur:
    cur.execute("select order_no, external_product_id from orders where platform='jd'")
    existing = {(str(order).strip(), str(sku).strip()) for order, sku in cur.fetchall()}
for row in rows:
    order, sku = str(row[1]).strip(), str(int(row[0]))
    print(order, sku, "IN_DB" if (order, sku) in existing else "MISSING")
