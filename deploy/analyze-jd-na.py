from openpyxl import load_workbook
from collections import Counter
import sys, json, re

path = sys.argv[1]
book = load_workbook(path, read_only=False, data_only=False)
sheet = book.active
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
print("HEADERS", list(enumerate(headers, 1)))
for row in sheet.iter_rows(min_row=2):
    values = [cell.value for cell in row]
    formulas = [(cell.column, cell.value) for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    if formulas: print("FORMULA", values[:2], formulas)

data = load_workbook(path, read_only=True, data_only=True).active
mapping = json.load(open(sys.argv[2], encoding="utf-8")) if len(sys.argv) > 2 else None
counts = Counter()
for row in data.iter_rows(min_row=2, values_only=True):
    if not any(value == "#N/A" or (isinstance(value, str) and "#N/A" in value) for value in row): continue
    if not mapping: print("NA", row); continue
    plan, sku, pin = str(row[13] or "").strip(), str(row[0] or "").split(".")[0], str(row[6] or "")
    ids = re.findall(r"\((\d+)\)", pin)
    reasons=[]
    if plan not in mapping['plans']: reasons.append('未命中计划')
    if sku not in mapping['skus']: reasons.append('未命中SKU')
    if not any(x in mapping['alliances'] for x in ids): reasons.append('未命中联盟ID')
    reason='、'.join(reasons) or '已通过规则，需核查导出范围/同步时间'
    counts[reason]+=1
    print('NA',row[1],sku,plan,pin,reason)
print('SUMMARY',counts)
